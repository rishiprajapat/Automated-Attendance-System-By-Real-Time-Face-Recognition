# import openpyxl and tkinter modules 
from openpyxl import *
from tkinter import *
import os
  
# globally declare wb and sheet variable 
  
# opening the existing excel file 
wb = load_workbook('data.xlsx') 
  
# create the sheet object 
sheet = wb.active 
    
# Function to set focus (cursor) 
def focus1(event): 
    # set focus on the course_field box 
    name_field.focus_set() 
def focus2(event): 
    # set focus on the course_field box 
    course_field.focus_set()


  
# Function for clearing the 
# contents of text entry boxes 
def clear(): 
      
    # clear the content of text entry box 
    name_field.delete(0, END) 
    course_field.delete(0, END) 
    
  
  
# Function to take data from GUI  
# window and write to an excel file 
def insert(): 
      
    # if user not fill any entry 
    # then print "empty input" 
    if (name_field.get() == "" and
        course_field.get() == ""): 
              
            print("empty input") 
  
    else: 
  
        # assigning the max row and max column 
        # value upto which data is written 
        # in an excel sheet to the variable 
        current_row = sheet.max_row 
        current_column = sheet.max_column 
  
        # get method returns current text 
        # as string which we write into 
        # excel spreadsheet at particular location 
        sheet.cell(row=current_row + 1, column=1).value = name_field.get() 
        sheet.cell(row=current_row + 1, column=2).value = int(course_field.get())
        # sheet.cell(row=current_row + 1, column=3).value = 'A'

        
        
       
  
        # save the file 
        wb.save('data.xlsx') 
  
        # set focus on the name_field box 
        name_field.focus_set() 
  
        # call the clear() function 
        clear() 


def function1():
    
    os.system("py train.py")




# Driver code 
if __name__ == "__main__": 
      
    # create a GUI window 
    root = Tk() 
  
    # set the background colour of GUI window 
    root.configure(background='light green') 
  
    # set the title of GUI window 
    root.title("registration form") 
  
    # set the configuration of GUI window 
    root.geometry("600x500") 
  
   
  
    # create a Form label 
    heading = Label(root, text="Form", bg="light green") 
  
    # create a Name label 
    name = Label(root, text="Name", bg="light green") 
  
    # create a Course label 
    course = Label(root, text="Enrollment no", bg="light green") 

  

  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    heading.grid(row=0, column=1) 
    name.grid(row=1, column=0) 
    course.grid(row=2, column=0) 
    
    
  
    # create a text entry box 
    # for typing the information 
    name_field = Entry(root) 
    course_field = Entry(root) 
  
    # bind method of widget is used for 
    # the binding the function with the events 
  
    # whenever the enter key is pressed 
    # then call the focus1 function 
    name_field.bind("<Return>", focus1) 
  
    # whenever the enter key is pressed 
    # then call the focus2 function 
    course_field.bind("<Return>", focus2) 

  
    
  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    name_field.grid(row=1, column=1, ipadx="100") 
    course_field.grid(row=2, column=1, ipadx="100") 

    # # call excel function 
    # excel() 
     
    # create a Submit Button and place into the root window 
    submit = Button(root, text="register student", fg="Black", 
                            bg="white", command=insert) 
    submit.grid(row=3, column=1) 
    
    Button(root,text="Train Model",font=("times new roman",10),bg="white",fg='black',command=function1).grid(row=4,columnspan=2,sticky=W+E+N+S,padx=100,pady=20)

   
    

    # start the GUI 
    root.mainloop() 
