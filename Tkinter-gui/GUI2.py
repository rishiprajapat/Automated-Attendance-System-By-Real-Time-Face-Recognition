from openpyxl import *
from tkinter import *
import os

# globally declare wb and sheet variable 

# opening the existing excel file 
wb = load_workbook('data.xlsx') 

# create the sheet object 
sheet = wb.active 


def excel(): 
    wb = load_workbook('data.xlsx') 
    current_row = sheet.max_row 
    current_column = sheet.max_column
    print(current_column)
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Enrollment"
    wb.save('data.xlsx') 
 

# Function to set focus (cursor) 
def focus1(event):  
	date_field.focus_set() 


def clear():
    date_field.delete(0, END) 

	 
def insert(): 
	if (date_field.get() == ""): 
			
		print("empty input") 

	else: 
	
		current_row = sheet.max_row 
		current_column = sheet.max_column 

		sheet.cell(row=1, column=current_column+1).value = date_field.get()
		
		# save the file 
		wb.save('data.xlsx') 

		# set focus on the name_field box 
		date_field.focus_set() 

		# call the clear() function 
		clear() 

def function1():
    
    os.system("py detect.py")
def function2():
    
    os.system("py update.py")

# Driver code 
if __name__ == "__main__": 
	
    # create a GUI window 
    root = Tk() 

    # set the background colour of GUI window 
    root.configure(background='light green') 

    # set the title of GUI window 
    root.title("registration form") 

    # set the configuration of GUI window 
    root.geometry("500x300") 

    # excel() 

    # # create a Form label 
    heading = Label(root, text="Form", bg="light green") 

    # # create a Name label 
    date = Label(root, text="Date", bg="light green") 

    heading.grid(row=0, column=1) 
    date.grid(row=1, column=0) 
   
    date_field = Entry(root,textvariable='info') 
   
    date_field.bind("<Return>", focus1) 

    date_field.grid(row=1, column=1, ipadx="100") 
    # course_field.grid(row=2, column=1, ipadx="100") 
    Button(root,text="Detect for Attendence",font=("times new roman",10),bg="white",fg='black',command=function1).grid(row=5,columnspan=2,sticky=W+E+N+S,padx=100,pady=20)

    Button(root,text="Update excel",font=("times new roman",10),bg="white",fg='black',command=function2).grid(row=6,columnspan=2,sticky=W+E+N+S,padx=100,pady=20)
    # call excel function 
    excel()
    

    

 




    # # create a Submit Button and place into the root window 
    submit = Button(root, text="Submit", fg="Black", 
                            bg="Red", command=insert) 
    submit.grid(row=8, column=1) 

    # # start the GUI 
    root.mainloop() 
